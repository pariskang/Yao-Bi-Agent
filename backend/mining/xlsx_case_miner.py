"""De-identifying xlsx case miner for Shen Qinrong lumbar-Bi outpatient records.

读取门诊 xlsx 导出（含姓名/病案号/地址等 PII），在内存中立即脱敏：
只保留性别、年龄段、症状关键词标签、证型、诊断、方药与剂量等研究字段；
姓名、病案号、地址、医师工号、就诊序号一律丢弃，原始自由文本不落盘、不导出。

挖掘输出全部是聚合统计（频次、support/confidence/lift、剂量分布、签名方剂命中），
证据引用只使用 xlsx 行号。所有候选规则 status=pending_expert_review，
仅供医师审核与科研教学，不构成诊断或处方依据。

CLI::

    python -m backend.mining.xlsx_case_miner --xlsx 病例.xlsx \
        --yaml rules/11_mined_rule_candidates.yaml \
        --frontend frontend/mined_rules.js
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

HERB_RE = re.compile(r"\d+/\s*\*?\s*([一-龥（）()]+?)\s*\*\s*\d+(?:\.\d+)?克/(\d+(?:\.\d+)?)克")
ZHENG_RE = re.compile(r"证型[:：]\s*([^,，/\s]+)")
AGE_RE = re.compile(r"(\d+)")
WESTERN_CODE_RE = re.compile(r"\([A-Z0-9]+\)T?")

PII_COLUMNS = {"姓名", "病案号", "地址", "医师工号", "就诊序号", "就诊日期", "医师姓名"}

# 主诉/现病史关键词 → 项目统一 normalized tag。
SYMPTOM_TAG_KEYWORDS: dict[str, list[str]] = {
    "lower_limb_numbness": ["麻木", "麻痛"],
    "radiating_leg_pain": ["放射", "窜痛", "牵及下肢"],
    "bilateral_leg_involvement": ["双下肢"],
    "cold_aggravation": ["遇冷", "受凉", "怕冷", "畏寒"],
    "cold_pain": ["冷痛"],
    "bitter_taste": ["口苦"],
    "insomnia": ["失眠", "寐差", "睡眠差", "夜寐"],
    "fatigue": ["乏力", "神疲"],
    "poor_appetite": ["纳差", "胃口差"],
    "distending_pain": ["胀痛", "酸胀"],
    "soreness": ["酸痛", "腰酸"],
    "activity_limitation": ["活动受限", "弯腰受限"],
    "night_pain": ["夜间痛", "夜间加重"],
    "sedentary_aggravation": ["久坐"],
    "acute_on_chronic": ["加重"],
}

# 功效药物模块（按沈氏处方实际用名归类，含炮制前缀变体）。
HERB_MODULES: dict[str, set[str]] = {
    "祛风除湿": {"独活", "秦艽", "防风", "桑寄生", "威灵仙", "羌活", "徐长卿"},
    "温阳散寒": {"附片", "肉桂", "细辛", "麻黄", "干姜", "桂枝"},
    "活血通络": {"当归", "川芎", "丹参", "赤芍", "燀山桃仁", "牡丹皮", "鸡血藤", "红花"},
    "补肝肾强筋骨": {"盐杜仲", "川牛膝", "熟地黄", "蒸萸肉", "盐补骨脂", "桑寄生", "续断", "狗脊", "淫羊藿"},
    "健脾化湿": {"白术", "茯苓", "薏苡仁", "炒薏苡仁", "陈皮", "党参", "炒党参", "姜半夏", "山药"},
    "和解少阳": {"北柴胡", "黄芩", "姜半夏"},
    "虫类搜络": {"全蝎", "蜈蚣", "地龙", "僵蚕", "乌梢蛇"},
    "益气固表": {"黄芪", "党参", "炒党参"},
    "养血柔筋": {"麸白芍", "当归", "熟地黄", "大枣"},
    "清热坚阴": {"知母", "黄芩", "牡丹皮"},
    "安神除烦": {"制远志肉", "炒酸枣仁", "首乌藤", "煅龙骨", "煅牡蛎"},
}

# 经典基础方签名（命中 ≥ SIGNATURE_THRESHOLD 比例的签名药物即记一次路线命中）。
FORMULA_SIGNATURES: dict[str, set[str]] = {
    "独活寄生汤": {"独活", "桑寄生", "秦艽", "防风", "细辛", "川芎", "熟地黄", "麸白芍", "肉桂", "茯苓", "盐杜仲", "川牛膝"},
    "当归四逆汤": {"当归", "桂枝", "麸白芍", "细辛", "通草", "大枣", "炙甘草"},
    "桂枝芍药知母汤": {"桂枝", "麸白芍", "知母", "麻黄", "附片", "白术", "防风"},
    "黄芪桂枝五物汤": {"黄芪", "桂枝", "麸白芍", "生姜", "大枣"},
    "小柴胡汤类(和解少阳)": {"北柴胡", "黄芩", "姜半夏"},
    "四物/八珍气血底盘": {"当归", "川芎", "熟地黄", "麸白芍", "党参", "白术", "茯苓"},
}
SIGNATURE_THRESHOLD = 0.7

# 需医师重点复核剂量的药物（毒性/猛药/超药典常量）。
DOSE_WATCHLIST = ["细辛", "附片", "全蝎", "蜈蚣", "麻黄", "制川乌", "制草乌", "白术", "黄芪"]

ASSOC_MIN_BOTH = 4
ASSOC_MIN_SUPPORT = 0.04
ASSOC_MIN_CONFIDENCE = 0.30
ASSOC_MIN_LIFT = 1.15


def parse_herbs(raw: str | None) -> list[tuple[str, float]]:
    if not raw:
        return []
    return [(m.group(1).strip(), float(m.group(2))) for m in HERB_RE.finditer(raw)]


def parse_zheng(raw: str | None) -> str | None:
    m = ZHENG_RE.search(raw or "")
    return m.group(1).strip() if m else None


def parse_tcm_diseases(raw: str | None) -> list[str]:
    text = (raw or "").replace("中医诊断:", "").split("/")[0]
    return [p.strip() for p in re.split(r"[,，;；]", text) if p.strip()]


def parse_western_dx(raw: str | None) -> list[str]:
    text = WESTERN_CODE_RE.sub("", (raw or "").replace("西医诊断:", ""))
    return [p.strip() for p in re.split(r"[,，;；]", text) if p.strip()]


def parse_age(raw: Any) -> int | None:
    m = AGE_RE.match(str(raw or ""))
    return int(m.group(1)) if m else None


def symptom_tags(text: str) -> list[str]:
    tags = [tag for tag, keys in SYMPTOM_TAG_KEYWORDS.items() if any(k in text for k in keys)]
    return sorted(tags)


def load_cases_from_xlsx(xlsx_path: str | Path, sheet: str | None = None) -> list[dict[str, Any]]:
    """Load and immediately de-identify outpatient rows.

    返回的 case dict 不包含任何 PII；自由文本仅做关键词扫描后丢弃。
    """

    import openpyxl  # optional dependency: pip install .[mining]

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}

    def col(row: tuple, name: str) -> Any:
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    cases: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(row):
            continue
        free_text = " ".join(str(col(row, c) or "") for c in ("主诉", "现病史", "既往史"))
        herbs = parse_herbs(col(row, "中药"))
        age = parse_age(col(row, "年龄"))
        case = {
            "row_id": row_number,  # 仅引用行号，不携带身份信息
            "sex": str(col(row, "性别") or "").strip() or None,
            "age": age,
            "age_band": f"{age // 10 * 10}s" if age is not None else None,
            "symptom_tags": symptom_tags(free_text),
            "zheng": parse_zheng(col(row, "中医诊断")),
            "tcm_diseases": parse_tcm_diseases(col(row, "中医诊断")),
            "western_dx": parse_western_dx(col(row, "西医诊断")),
            "herbs": [name for name, _ in herbs],
            "herb_doses": {name: dose for name, dose in herbs},
            "has_prescription": bool(herbs),
            "osteoporosis_history": "骨质疏松" in free_text or any("骨质疏松" in d for d in parse_western_dx(col(row, "西医诊断"))),
        }
        if case["age"] is not None and case["age"] >= 60:
            case["symptom_tags"] = sorted(set(case["symptom_tags"]) | {"elderly"})
        cases.append(case)
    wb.close()
    return cases


# ---------------------------------------------------------------------------
# Mining
# ---------------------------------------------------------------------------


def _herb_modules_of(herbs: set[str]) -> list[str]:
    return sorted(module for module, members in HERB_MODULES.items() if len(herbs & members) >= 2 or (len(members & herbs) >= 1 and module == "虫类搜络"))


def _formula_hits(herbs: set[str]) -> list[str]:
    hits = []
    for name, signature in FORMULA_SIGNATURES.items():
        if len(herbs & signature) / len(signature) >= SIGNATURE_THRESHOLD:
            hits.append(name)
    return hits


def _association(transactions: list[set[str]], antecedent: str, consequent: str) -> dict[str, Any] | None:
    n = len(transactions)
    if not n:
        return None
    has_a = [t for t in transactions if antecedent in t]
    has_c = sum(1 for t in transactions if consequent in t)
    both = sum(1 for t in has_a if consequent in t)
    if both < ASSOC_MIN_BOTH or not has_a or not has_c:
        return None
    support = both / n
    confidence = both / len(has_a)
    lift = confidence / (has_c / n)
    if support < ASSOC_MIN_SUPPORT or confidence < ASSOC_MIN_CONFIDENCE or lift < ASSOC_MIN_LIFT:
        return None
    return {
        "antecedent": antecedent,
        "consequent": consequent,
        "support": round(support, 3),
        "confidence": round(confidence, 3),
        "lift": round(lift, 2),
        "n_both": both,
        "n_antecedent": len(has_a),
    }


def mine_cases(cases: list[dict[str, Any]]) -> dict[str, Any]:
    rx_cases = [c for c in cases if c["has_prescription"]]
    zheng_dist = Counter(c["zheng"] or "未标证型" for c in cases)
    sex_dist = Counter(c["sex"] or "未知" for c in cases)
    age_band_dist = Counter(c["age_band"] or "未知" for c in cases)
    symptom_dist = Counter(tag for c in cases for tag in c["symptom_tags"])
    western_dist = Counter(dx for c in cases for dx in c["western_dx"])

    herb_freq = Counter(h for c in rx_cases for h in set(c["herbs"]))
    formula_counter: Counter[str] = Counter()
    formula_zheng: dict[str, Counter] = defaultdict(Counter)
    formula_rows: dict[str, list[int]] = defaultdict(list)
    module_counter: Counter[str] = Counter()
    transactions: list[set[str]] = []
    for c in rx_cases:
        herbs = set(c["herbs"])
        modules = _herb_modules_of(herbs)
        formulas = _formula_hits(herbs)
        module_counter.update(modules)
        for f in formulas:
            formula_counter[f] += 1
            formula_zheng[f][c["zheng"] or "未标证型"] += 1
            formula_rows[f].append(c["row_id"])
        items = set(c["symptom_tags"])
        items.update(f"module::{m}" for m in modules)
        items.update(f"formula::{f}" for f in formulas)
        if c["zheng"]:
            items.add(f"zheng::{c['zheng']}")
        if c["osteoporosis_history"]:
            items.add("osteoporosis")
        transactions.append(items)

    associations: list[dict[str, Any]] = []
    consequents = [f"formula::{f}" for f in FORMULA_SIGNATURES] + [f"module::{m}" for m in HERB_MODULES]
    antecedents = sorted({t for tr in transactions for t in tr if not t.startswith(("formula::", "module::"))})
    for a in antecedents:
        for cq in consequents:
            found = _association(transactions, a, cq)
            if found:
                associations.append(found)
    associations.sort(key=lambda r: (r["lift"], r["confidence"]), reverse=True)

    dose_table = {}
    for herb in DOSE_WATCHLIST:
        doses = [c["herb_doses"][herb] for c in rx_cases if herb in c["herb_doses"]]
        if doses:
            counts = Counter(doses)
            dose_table[herb] = {
                "n": len(doses),
                "min_g": min(doses),
                "max_g": max(doses),
                "mode_g": counts.most_common(1)[0][0],
                "clinician_review_required": True,
            }

    data_quality = {
        "n_cases": len(cases),
        "n_with_prescription": len(rx_cases),
        "n_missing_zheng": sum(1 for c in cases if not c["zheng"]),
        "tongue_pulse_usable": False,
        "tongue_pulse_note": "中医四诊栏几乎全部为模板文本，舌脉信息无法用于规则挖掘，证候规则只能依靠症状关键词与诊断字段。",
    }

    return {
        "dataset_stats": {
            "n_cases": len(cases),
            "n_with_prescription": len(rx_cases),
            "sex_distribution": dict(sex_dist),
            "age_band_distribution": dict(sorted(age_band_dist.items())),
            "zheng_distribution": dict(zheng_dist.most_common()),
            "western_dx_top": dict(western_dist.most_common(12)),
            "symptom_tag_distribution": dict(symptom_dist.most_common()),
        },
        "herb_frequency_top": dict(herb_freq.most_common(40)),
        "herb_module_counts": dict(module_counter.most_common()),
        "formula_signature_hits": [
            {
                "formula": f,
                "n_cases": n,
                "by_zheng": dict(formula_zheng[f].most_common()),
                "evidence_rows": formula_rows[f][:30],
            }
            for f, n in formula_counter.most_common()
        ],
        "associations": associations,
        "dose_table": dose_table,
        "data_quality": data_quality,
    }


# ---------------------------------------------------------------------------
# Rule candidates
# ---------------------------------------------------------------------------


def build_rule_candidates(mined: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert aggregates into reviewable rule candidates (never auto-enabled)."""

    candidates: list[dict[str, Any]] = []
    seq = 0

    def add(rule_type: str, condition: dict[str, Any], conclusion: dict[str, Any], stats: dict[str, Any], note: str) -> None:
        nonlocal seq
        seq += 1
        candidates.append({
            "rule_id": f"MINED-{rule_type.upper()}-{seq:03d}",
            "rule_type": rule_type,
            "if": condition,
            "then": conclusion,
            "statistics": stats,
            "evidence": note,
            "source": "xlsx_outpatient_mining_deidentified",
            "status": "pending_expert_review",
            "clinician_only": True,
            "strength": "weak_statistical_signal" if stats.get("lift", 0) < 2 else "moderate_statistical_signal",
        })

    n_rx = mined["dataset_stats"]["n_with_prescription"]
    for hit in mined["formula_signature_hits"]:
        top_zheng = next(iter(hit["by_zheng"]), None)
        add(
            "formula_route",
            {"zheng_any": list(hit["by_zheng"])[:3]},
            {"candidate_formula_route": hit["formula"]},
            {"n_cases": hit["n_cases"], "support": round(hit["n_cases"] / n_rx, 3) if n_rx else 0, "top_zheng": top_zheng},
            f"{hit['n_cases']}/{n_rx} 张处方命中该签名；证据行号（脱敏）：{hit['evidence_rows'][:10]}",
        )

    for assoc in mined["associations"]:
        target_kind = "formula" if assoc["consequent"].startswith("formula::") else "module"
        add(
            f"{target_kind}_association",
            {"tag": assoc["antecedent"]},
            {f"candidate_{target_kind}": assoc["consequent"].split("::", 1)[1]},
            {k: assoc[k] for k in ("support", "confidence", "lift", "n_both", "n_antecedent")},
            f"标签 {assoc['antecedent']} 与 {assoc['consequent']} 共现 {assoc['n_both']} 例。",
        )

    for herb, stats in mined["dose_table"].items():
        add(
            "dose_convention",
            {"herb": herb},
            {"observed_dose_range_g": [stats["min_g"], stats["max_g"]], "mode_dose_g": stats["mode_g"]},
            {"n": stats["n"]},
            "仅作为医师端经验研究剂量分布，不向患者输出，不构成可执行剂量。",
        )
    return candidates


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------


def write_outputs(
    mined: dict[str, Any],
    candidates: list[dict[str, Any]],
    yaml_path: str | Path | None = None,
    frontend_path: str | Path | None = None,
) -> dict[str, str]:
    import yaml

    written: dict[str, str] = {}
    meta = {
        "generated_on": date.today().isoformat(),
        "generator": "backend.mining.xlsx_case_miner",
        "privacy": "de-identified aggregates only; evidence references are xlsx row numbers",
        "review_policy": "all candidates pending_expert_review; clinician-only; never patient-executable",
    }
    if yaml_path:
        payload = {"meta": meta, "dataset_stats": mined["dataset_stats"], "dose_table": mined["dose_table"], "data_quality": mined["data_quality"], "rule_candidates": candidates}
        Path(yaml_path).write_text(yaml.safe_dump(payload, allow_unicode=True, sort_keys=False), encoding="utf-8")
        written["yaml"] = str(yaml_path)
    if frontend_path:
        js_payload = {"meta": meta, **mined, "rule_candidates": candidates}
        Path(frontend_path).write_text(
            "// Auto-generated by backend.mining.xlsx_case_miner — de-identified aggregates only.\n"
            "window.MINED_RULES = " + json.dumps(js_payload, ensure_ascii=False, indent=2) + ";\n",
            encoding="utf-8",
        )
        written["frontend"] = str(frontend_path)
    return written


def main() -> None:
    parser = argparse.ArgumentParser(description="De-identify and mine Shen Qinrong lumbar-Bi xlsx cases into rule candidates")
    parser.add_argument("--xlsx", required=True, help="Path to the raw outpatient xlsx (kept local, never committed)")
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--yaml", default="rules/11_mined_rule_candidates.yaml")
    parser.add_argument("--frontend", default="frontend/mined_rules.js")
    parser.add_argument("--print-stats", action="store_true")
    args = parser.parse_args()
    cases = load_cases_from_xlsx(args.xlsx, sheet=args.sheet)
    mined = mine_cases(cases)
    candidates = build_rule_candidates(mined)
    written = write_outputs(mined, candidates, yaml_path=args.yaml, frontend_path=args.frontend)
    if args.print_stats:
        print(json.dumps(mined["dataset_stats"], ensure_ascii=False, indent=2))
    print(f"cases={len(cases)} candidates={len(candidates)} written={written}")


if __name__ == "__main__":
    main()
